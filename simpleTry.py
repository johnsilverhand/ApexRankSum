import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def initialize_excel_file(input_file_path, output_file_path):
    df = pd.read_excel(input_file_path)
    df['实力分'] = df['段位分'] + df['场均伤害分']
    df['总积分'] = 0
    df.loc[df['实力分'] > 20, '总积分'] = -2 * (df['实力分'] - 20)
    df.loc[df['实力分'] < 14, '总积分'] = 2 * (14 - df['实力分'])
    df['比赛积分'] = 0
    df['赛点'] = '否'
    df.to_excel(output_file_path, index=False)

def update_team_strength(file_path):
    df = pd.read_excel(file_path)
    print("Available Teams:")
    for i, team in enumerate(df['队伍名'], start=1):
        print(f"{i}. {team}")
    team_idx = int(input("请输入你想更新实力分的队伍的编号: ")) - 1
    if team_idx not in range(len(df)):
        print("Invalid team number. Exiting.")
        return
    new_段位分 = float(input("请输入新段位分: "))
    new_场均伤害分 = float(input("请输入新的场均伤害分: "))
    df.loc[team_idx, '段位分'] = new_段位分
    df.loc[team_idx, '场均伤害分'] = new_场均伤害分
    df.loc[team_idx, '实力分'] = new_段位分 + new_场均伤害分
    df.loc[team_idx, '总积分'] = df.loc[team_idx, '比赛积分'] + df.loc[team_idx, '实力分']
    df.to_excel(file_path, index=False)

def settle_single_round(start_file, single_round_file, settlement_file):
    df_start = pd.read_excel(start_file)
    df_single_round = pd.read_excel(single_round_file)
    rank_points = {1: 12, 2: 9, 3: 7, 4: 5, 5: 4, 6: 3, 7: 3, 8: 2, 9: 2, 10: 2, 11: 1, 12: 1, 13: 1, 14: 1, 15: 1}
    df_settlement = pd.DataFrame(columns=["队伍名", "人头分", "排名分", "本局得分",'排名','击杀人数'])
    for index, row in df_single_round.iterrows():
        team_name = row['队伍名']
        kills = min(row['击杀人数'], 6)
        rank = row['排名']
        rank_point = rank_points.get(rank, 0)
        total_points = kills + rank_point
        df_start.loc[df_start['队伍名'] == team_name, '比赛积分'] += total_points
        df_start.loc[df_start['队伍名'] == team_name, '总积分'] = df_start.loc[df_start['队伍名'] == team_name, '比赛积分'] + df_start.loc[df_start['队伍名'] == team_name, '实力分']
        if df_start.loc[df_start['队伍名'] == team_name, '总积分'].iloc[0] >= 45:
            df_start.loc[df_start['队伍名'] == team_name, '赛点'] = '是'
        df_settlement = df_settlement.append({"队伍名": team_name, "人头分": kills, "排名分": rank_point, "本局得分": total_points,"排名":rank,"击杀人数":row['击杀人数']}, ignore_index=True)
    df_start.sort_values(by="总积分", ascending=False, inplace=True)
    df_settlement.sort_values(by="本局得分", ascending=False, inplace=True)
    df_start.to_excel(start_file, index=False)
    df_settlement.to_excel(settlement_file, index=False)
    book = load_workbook(start_file)
    sheet = book.active
    red_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in range(2, sheet.max_row + 1):
        if sheet[f'G{row}'].value == '是':
            for col in ['A', 'F', 'G']:
                sheet[f'{col}{row}'].fill = red_fill
    book.save(start_file)

def main():
    while True:
        print("1. 比赛初始化")
        print("2. 实力分更新")
        print("3. 单局结算")
        print("4. 退出")
        choice = input("输入序号执行对应功能: ")
        if choice == '1':
            input_file_path = "start.xlsx"
            output_file_path = "start.xlsx"
            initialize_excel_file(input_file_path, output_file_path)
        elif choice == '2':
            file_path = "start.xlsx"
            update_team_strength(file_path)
        elif choice == '3':
            start_file = "start.xlsx"
            single_round_file = "single_round.xlsx"
            settlement_file = "单局结算.xlsx"
            settle_single_round(start_file, single_round_file, settlement_file)
        elif choice == '4':
            break
        else:
            print("Invalid choice. Please enter a number between 1 and 4.")

if __name__ == "__main__":
    main()

